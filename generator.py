"""
Phase 3 — Générateur Excel + Rapport TXT (version minimaliste)
Tableau brut + orange/rouge pour les alertes + gestion des fusions
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def generer_excel(resultat: dict, image_path: str) -> Workbook:
    """
    Génère un Excel avec le tableau brut + couleurs uniquement sur les alertes
    Gère les cellules fusionnées (colspan/rowspan)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction"

    colonnes = resultat.get("colonnes", [])
    lignes = resultat.get("lignes", [])
    
    # Récupérer les noms des colonnes
    noms_colonnes = [col.get("nom", col.get("valeur", "")) for col in colonnes]

    # Styles : uniquement orange et rouge
    ORANGE = PatternFill("solid", start_color="FFEB9C")
    ROUGE = PatternFill("solid", start_color="FFC7CE")
    FONT_NORMAL = Font(name="Arial", size=10)
    FONT_BOLD = Font(name="Arial", size=10, bold=True)

    # --- Ligne 1 : En-têtes ---
    for j, nom in enumerate(noms_colonnes):
        cell = ws.cell(row=1, column=j + 1)
        cell.value = nom
        cell.font = FONT_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Parcours des lignes pour détecter et appliquer les fusions ---
    matrix = {}
    
    # Remplir la matrice avec les données
    for i, ligne in enumerate(lignes):
        row_idx = i + 2
        for j, nom_col in enumerate(noms_colonnes):
            col_idx = j + 1
            cellule = ligne.get(nom_col, {})
            
            valeur = cellule.get("valeur", "")
            statut = cellule.get("statut", "vert")
            rowspan = cellule.get("rowspan", 1)
            colspan = cellule.get("colspan", 1)
            
            matrix[(row_idx, col_idx)] = {
                "valeur": valeur,
                "statut": statut,
                "rowspan": rowspan,
                "colspan": colspan,
                "est_fusion": rowspan > 1 or colspan > 1
            }
    
    # Appliquer les fusions et écrire les valeurs
    fusions_appliquees = set()
    
    for (row, col), data in matrix.items():
        if (row, col) in fusions_appliquees:
            continue
            
        rowspan = data["rowspan"]
        colspan = data["colspan"]
        
        cell = ws.cell(row=row, column=col)
        
        # Conversion en nombre si possible
        valeur = data["valeur"]
        if valeur and isinstance(valeur, str):
            try:
                valeur = float(valeur.replace(',', '.'))
            except:
                pass
        
        cell.value = valeur
        cell.font = FONT_NORMAL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Appliquer la couleur si nécessaire
        if data["statut"] == "orange":
            cell.fill = ORANGE
        elif data["statut"] == "rouge":
            cell.fill = ROUGE
        
        # Appliquer la fusion si besoin
        if rowspan > 1 or colspan > 1:
            end_row = row + rowspan - 1
            end_col = col + colspan - 1
            
            if end_row > row or end_col > col:
                ws.merge_cells(
                    start_row=row, start_column=col,
                    end_row=end_row, end_column=end_col
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for r in range(row, end_row + 1):
                for c in range(col, end_col + 1):
                    fusions_appliquees.add((r, c))
    
    # --- Ajustement largeur colonnes ---
    for j in range(1, len(noms_colonnes) + 1):
        col_letter = ws.cell(row=1, column=j).column_letter
        max_len = len(str(noms_colonnes[j-1]))
        for i in range(2, len(lignes) + 2):
            val = ws.cell(row=i, column=j).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    return wb


def generer_rapport_txt(resultat: dict, image_path: str) -> str:
    """
    Génère un rapport TXT avec uniquement les alertes (orange et rouge)
    """
    rapport = resultat.get("rapport", {})
    lignes = resultat.get("lignes", [])
    colonnes = resultat.get("colonnes", [])
    
    noms_colonnes = [col.get("nom", col.get("valeur", "")) for col in colonnes]
    
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    lignes_txt = [
        "=" * 60,
        "RAPPORT DE CONTRÔLE QUALITÉ",
        "=" * 60,
        f"Document  : {Path(image_path).name}",
        f"Titre     : {resultat.get('titre', {}).get('valeur', '')}",
        f"Date      : {now}",
        "",
        "RÉSUMÉ",
        "-" * 40,
        f"Score global     : {rapport.get('score_global', 0)}%",
        f"Cellules OK      : {rapport.get('verts', 0)}",
        f"Alertes orange   : {rapport.get('oranges', 0)}",
        f"Alertes rouges   : {rapport.get('rouges', 0)}",
        f"Total cellules   : {rapport.get('total_cellules', 0)}",
        "",
    ]

    # ALERTES - Titre
    t = resultat.get("titre", {})
    if t.get("statut") != "vert":
        lignes_txt += [
            "⚠️ ALERTE — TITRE",
            "-" * 40,
            f"  Statut   : {t.get('statut', '').upper()}",
            f"  Valeur A : {t.get('valeur', '')}",
            f"  Valeur B : {t.get('valeur_b', '')}",
            f"  Confiance: A={t.get('confiance_a', 0)} | B={t.get('confiance_b', 0)}",
            "",
        ]

    # ALERTES - En-têtes
    alertes_entetes = [col for col in colonnes if col.get("statut") != "vert"]
    if alertes_entetes:
        lignes_txt += ["⚠️ ALERTES — EN-TÊTES DE COLONNES", "-" * 40]
        for col in alertes_entetes:
            lignes_txt += [
                f"  Colonne  : {col.get('nom', col.get('valeur', ''))}",
                f"  Statut   : {col.get('statut', '').upper()}",
                f"  Valeur B : {col.get('valeur_b', '')}",
                f"  Confiance: A={col.get('confiance_a', 0)} | B={col.get('confiance_b', 0)}",
                "",
            ]

    # ALERTES - Données
    alertes_data = []
    for i, ligne in enumerate(lignes):
        for nom_col in noms_colonnes:
            cell = ligne.get(nom_col)
            if cell and cell.get("statut") != "vert":
                alertes_data.append((i + 1, nom_col, cell))

    if alertes_data:
        lignes_txt += ["⚠️ ALERTES — DONNÉES", "-" * 40]
        for num_ligne, nom_col, cell in alertes_data:
            lignes_txt += [
                f"  Ligne {num_ligne} | {nom_col}",
                f"  Statut   : {cell.get('statut', '').upper()}",
                f"  Valeur A : {cell.get('valeur', '')}",
                f"  Valeur B : {cell.get('valeur_b', '')}",
                f"  Confiance: A={cell.get('confiance_a', 0)} | B={cell.get('confiance_b', 0)}",
                "",
            ]

    if not alertes_entetes and not alertes_data and t.get("statut") == "vert":
        lignes_txt.append("✅ Aucune alerte — toutes les cellules sont fiables.")

    lignes_txt += ["=" * 60]
    return "\n".join(lignes_txt)


def sauvegarder(resultat: dict, image_path: str, dossier_sortie: str = "."):
    """
    Sauvegarde l'Excel et le rapport TXT
    """
    base = Path(dossier_sortie) / Path(image_path).stem
    
    # Excel
    wb = generer_excel(resultat, image_path)
    path_excel = f"{base}_extraction.xlsx"
    wb.save(path_excel)
    
    # Rapport TXT
    rapport_txt = generer_rapport_txt(resultat, image_path)
    path_txt = f"{base}_rapport.txt"
    with open(path_txt, "w", encoding="utf-8") as f:
        f.write(rapport_txt)
    
    return path_excel, path_txt
